"""Tabular interconversion: csv <-> tsv <-> xlsx.

XLSX is read/written via openpyxl in read_only/write_only mode for speed.
For multi-sheet xlsx inputs, only the active (or first) sheet is exported to
csv/tsv; full multi-sheet round-trips are out of scope.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import ClassVar, Iterator

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from typing_extensions import override

from convertfs.converter import Converter


class TabularConverter(Converter):
    INPUTS = (re.compile(r'^(.*)\.(csv|tsv|xlsx)$'),)
    OUTPUT_FILES = (
        Path('{}.csv'),
        Path('{}.tsv'),
        Path('{}.xlsx'),
    )

    _DELIMITERS: ClassVar[dict[str, str]] = {'csv': ',', 'tsv': '\t'}

    @override
    def process(self, source: Path, requested: Path) -> bytes:
        src_ext = source.suffix.lstrip('.').lower()
        out_ext = requested.suffix.lstrip('.').lower()

        rows = list(self._read(source, src_ext))
        return self._write(rows, out_ext)

    def _read(self, source: Path, ext: str) -> Iterator[list[str]]:
        if ext in self._DELIMITERS:
            with source.open('r', encoding='utf-8', newline='') as f:
                reader = csv.reader(f, delimiter=self._DELIMITERS[ext])
                yield from reader
            return
        if ext == 'xlsx':
            try:
                wb = openpyxl.load_workbook(
                    str(source), read_only=True, data_only=True
                )
            except InvalidFileException as exc:
                msg = f'Invalid xlsx file: {source.name}'
                raise ValueError(msg) from exc
            try:
                ws = wb.active or wb.worksheets[0]
                for row in ws.iter_rows(values_only=True):
                    yield ['' if cell is None else str(cell) for cell in row]
            finally:
                wb.close()
            return
        msg = f'Unsupported tabular input: {ext}'
        raise ValueError(msg)

    def _write(self, rows: list[list[str]], ext: str) -> bytes:
        if ext in self._DELIMITERS:
            buf = io.StringIO()
            writer = csv.writer(
                buf, delimiter=self._DELIMITERS[ext], lineterminator='\n'
            )
            writer.writerows(rows)
            return buf.getvalue().encode('utf-8')
        if ext == 'xlsx':
            wb = openpyxl.Workbook(write_only=True)
            ws = wb.create_sheet()
            for row in rows:
                ws.append(row)
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        msg = f'Unsupported tabular output: {ext}'
        raise ValueError(msg)
