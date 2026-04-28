from __future__ import annotations

import csv
import io
from pathlib import Path

import openpyxl
import pytest

from convertfs.converters.tabular import TabularConverter


def _write_csv(path: Path, rows: list[list[str]]) -> Path:
	buf = io.StringIO()
	csv.writer(buf, lineterminator='\n').writerows(rows)
	path.write_text(buf.getvalue(), encoding='utf-8')
	return path


def _write_xlsx(path: Path, rows: list[list[str]]) -> Path:
	wb = openpyxl.Workbook(write_only=True)
	ws = wb.create_sheet()
	for row in rows:
		ws.append(row)
	wb.save(str(path))
	return path


def _read_xlsx_rows(data: bytes) -> list[list[str]]:
	wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
	try:
		ws = wb.active or wb.worksheets[0]
		return [
			['' if cell is None else str(cell) for cell in row]
			for row in ws.iter_rows(values_only=True)
		]
	finally:
		wb.close()


def test_tabular_csv_to_tsv_swaps_delimiter(tmp_path: Path, convert_bytes) -> None:
	source = _write_csv(tmp_path / 'data.csv', [['a', 'b'], ['1', '2']])

	output = convert_bytes(TabularConverter(), source, tmp_path / 'data.tsv')

	text = output.decode('utf-8')
	assert text == 'a\tb\n1\t2\n'


def test_tabular_tsv_to_csv_swaps_delimiter(tmp_path: Path, convert_bytes) -> None:
	source = tmp_path / 'data.tsv'
	source.write_text('a\tb\n1\t2\n', encoding='utf-8')

	output = convert_bytes(TabularConverter(), source, tmp_path / 'data.csv')

	# csv.writer quotes only when needed; this payload has no commas so
	# the output should be plain.
	assert output.decode('utf-8') == 'a,b\n1,2\n'


def test_tabular_csv_to_xlsx_round_trips_via_openpyxl(tmp_path: Path, convert_bytes) -> None:
	rows = [['header1', 'header2'], ['x', 'y'], ['1', '2']]
	source = _write_csv(tmp_path / 'data.csv', rows)

	output = convert_bytes(TabularConverter(), source, tmp_path / 'data.xlsx')

	# xlsx files are zip archives starting with "PK".
	assert output[:2] == b'PK'
	assert _read_xlsx_rows(output) == rows


def test_tabular_xlsx_to_csv_extracts_active_sheet(tmp_path: Path, convert_bytes) -> None:
	source = _write_xlsx(tmp_path / 'data.xlsx', [['a', 'b'], ['1', '2']])

	output = convert_bytes(TabularConverter(), source, tmp_path / 'data.csv')

	assert output.decode('utf-8') == 'a,b\n1,2\n'


def test_tabular_csv_to_csv_is_no_op_passthrough(tmp_path: Path, convert_bytes) -> None:
	# Same-format short-circuit: source bytes returned verbatim.
	original = b'a,b\nthis,that\n'
	source = tmp_path / 'data.csv'
	source.write_bytes(original)

	output = convert_bytes(TabularConverter(), source, tmp_path / 'data.csv')

	assert output == original


def test_tabular_xlsx_to_xlsx_is_no_op_passthrough(tmp_path: Path, convert_bytes) -> None:
	# Critically, the round-trip via openpyxl loses formulas and multi-sheet
	# layout — the no-op path side-steps that loss.
	source = _write_xlsx(tmp_path / 'data.xlsx', [['a', 'b'], ['1', '2']])
	original = source.read_bytes()

	output = convert_bytes(TabularConverter(), source, tmp_path / 'data.xlsx')

	assert output == original


def test_tabular_csv_handles_quoted_fields(tmp_path: Path, convert_bytes) -> None:
	source = tmp_path / 'data.csv'
	source.write_text('name,note\nalice,"hello, world"\n', encoding='utf-8')

	output = convert_bytes(TabularConverter(), source, tmp_path / 'data.tsv')

	# The comma inside the quoted value must not become a column break.
	lines = output.decode('utf-8').splitlines()
	assert lines == ['name\tnote', 'alice\thello, world']


def test_tabular_handles_empty_input(tmp_path: Path, convert_bytes) -> None:
	source = tmp_path / 'empty.csv'
	source.write_text('', encoding='utf-8')

	output = convert_bytes(TabularConverter(), source, tmp_path / 'empty.tsv')

	assert output == b''


def test_tabular_xlsx_to_xlsx_passthrough_preserves_multiple_sheets(
	tmp_path: Path,
	convert_bytes,
) -> None:
	# Build a workbook with two sheets via the regular (non-write_only)
	# Workbook so multi-sheet semantics are clearly preserved.
	source = tmp_path / 'multi.xlsx'
	wb = openpyxl.Workbook()
	wb.active.title = 'first'
	wb.active['A1'] = 'one'
	second = wb.create_sheet('second')
	second['A1'] = 'two'
	wb.save(str(source))
	original = source.read_bytes()

	output = convert_bytes(TabularConverter(), source, tmp_path / 'multi.xlsx')

	assert output == original
	# Sanity-check that the no-op preserved both sheets.
	rt = openpyxl.load_workbook(io.BytesIO(output))
	assert rt.sheetnames == ['first', 'second']


def test_tabular_rejects_unsupported_output(tmp_path: Path) -> None:
	source = _write_csv(tmp_path / 'data.csv', [['a']])

	with pytest.raises(ValueError, match='Unsupported tabular output'):
		TabularConverter().process(source, tmp_path / 'data.bogus', tmp_path / "_dest.bin")
